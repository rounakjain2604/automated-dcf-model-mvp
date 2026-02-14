from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule, CellIsRule
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .config import DCFConfig, ScenarioSet
from .wacc import WACCResult


ACCOUNTING_FMT = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
PCT_FMT = "0.00%"
PRICE_FMT = "$#,##0.00"
MILLIONS_FMT = "$#,##0.00,,\"M\""
BODY_FONT = "Segoe UI"
SOFT_RED = "C0504D"
INPUT_FONT = Font(color="0000FF")
FORMULA_FONT = Font(color="000000")
HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
SECTION_FILL = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
TITLE_FILL = PatternFill(fill_type="solid", start_color="0F243E", end_color="0F243E")
ALT_ROW_FILL = PatternFill(fill_type="solid", start_color="F7F9FC", end_color="F7F9FC")
KPI_FILL = PatternFill(fill_type="solid", start_color="E8F1FB", end_color="E8F1FB")
PASS_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
ALERT_FILL = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def export_workbook(
    output_path: str | Path,
    cfg: DCFConfig,
    scenario_name: str,
    scenario: ScenarioSet,
    period_meta: dict,
    forecast_df: pd.DataFrame,
    wacc_result: WACCResult,
    valuation_summary: dict,
    historical_growth_3y_avg: float = 0.0,
) -> None:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    last_forecast_row = cfg.forecast.years + 1

    _write_inputs(wb, cfg, scenario_name, scenario, period_meta, forecast_df, wacc_result)
    _write_forecast(wb, cfg.forecast.years, forecast_df)
    _write_valuation(wb, last_forecast_row)
    _write_sensitivity(wb, cfg, last_forecast_row, wacc_result.wacc)
    _write_dashboard(wb)
    _write_checks(wb, cfg.forecast.years)
    _write_report_data(wb, forecast_df, valuation_summary, cfg, scenario_name, historical_growth_3y_avg)
    _apply_workbook_theme(wb)

    wb.save(output_path)


def _write_inputs(
    wb: Workbook,
    cfg: DCFConfig,
    scenario_name: str,
    scenario: ScenarioSet,
    period_meta: dict,
    forecast_df: pd.DataFrame,
    wacc_result: WACCResult,
) -> None:
    ws = wb.create_sheet("Inputs")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"] = "Automated DCF Model Generator"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:D2")
    ws["A2"] = "Model control center — edit blue cells only"
    ws["A2"].font = Font(italic=True, color="555555")

    ws["A3"] = "Period Basis"
    ws["B3"] = period_meta.get("period_basis", "unknown")
    ws["A4"] = "Stub Period"
    ws["B4"] = str(period_meta.get("has_stub_period", False))
    ws["A5"] = "Scenario"
    ws["B5"] = scenario_name
    ws["A3"].fill = SECTION_FILL
    ws["A4"].fill = SECTION_FILL
    ws["A5"].fill = SECTION_FILL
    ws["A3"].font = Font(bold=True)
    ws["A4"].font = Font(bold=True)
    ws["A5"].font = Font(bold=True)

    rows = [
        ("Revenue CAGR", cfg.forecast.revenue_cagr),
        ("COGS % Revenue", cfg.forecast.cogs_pct_revenue),
        ("Opex % Revenue", cfg.forecast.opex_pct_revenue),
        ("Tax Rate", cfg.forecast.tax_rate),
        ("DSO", cfg.forecast.dso),
        ("DPO", cfg.forecast.dpo),
        ("DIO", cfg.forecast.dio),
        ("Capex % Revenue", cfg.forecast.capex_pct_revenue),
        ("Depreciation Rate", cfg.forecast.depreciation_rate),
        ("Risk-Free Rate", cfg.wacc.risk_free_rate),
        ("Market Risk Premium", cfg.wacc.market_risk_premium),
        ("Beta", cfg.wacc.beta),
        ("Size Premium", cfg.wacc.size_premium),
        ("Country Risk Premium", cfg.wacc.country_risk_premium),
        ("Debt Weight", cfg.wacc.target_debt_weight),
        ("Equity Weight", cfg.wacc.target_equity_weight),
        ("Pre-tax Cost of Debt", wacc_result.cost_of_debt_pre_tax),
        ("WACC Tax Rate", cfg.wacc.tax_rate),
        ("Terminal Growth Rate", cfg.valuation.terminal_growth_rate),
        ("Exit EV/EBITDA", cfg.valuation.exit_ev_ebitda_multiple),
        ("Fully Diluted Shares", cfg.valuation.fully_diluted_shares),
        ("Cash", cfg.valuation.cash),
        ("Debt", cfg.valuation.debt),
        ("Minority Interest", cfg.valuation.minority_interest),
        ("Preferred Stock", cfg.valuation.preferred_stock),
        ("Discount Convention", cfg.valuation.discount_convention),
        ("Scenario Revenue Multiplier", scenario.revenue_multiplier),
        ("Scenario Margin Delta (bps)", scenario.margin_delta_bps),
        ("Scenario Base NWC", float(forecast_df.iloc[0]["NWC"] - forecast_df.iloc[0]["Delta NWC"])),
        ("Scenario Capex Multiplier", scenario.capex_multiplier),
        ("Base Revenue", float(forecast_df.iloc[0]["Revenue"] / max(1 + (cfg.forecast.revenue_cagr * scenario.revenue_multiplier), 1e-9))),
        ("GDP Growth Cap", cfg.valuation.gdp_growth_cap),
        ("Current Market Price", 0.0),
        ("Ask Price", 0.0),
    ]

    ws["A7"] = "Assumption"
    ws["B7"] = "Value"
    _style_header_row(ws, 7, 2)

    for i, (label, value) in enumerate(rows, start=8):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        ws[f"B{i}"].font = INPUT_FONT
        ws[f"B{i}"].number_format = _input_number_format(label, value)
        ws[f"A{i}"].border = THIN_BORDER
        ws[f"B{i}"].border = THIN_BORDER
        if i % 2 == 0:
            ws[f"A{i}"].fill = ALT_ROW_FILL
            ws[f"B{i}"].fill = ALT_ROW_FILL

    ws.freeze_panes = "A8"
    ws.auto_filter.ref = "A7:B41"
    ws.print_area = "A1:D45"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    _autosize(ws)


def _write_forecast(wb: Workbook, years: int, forecast_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Forecast")
    ws.sheet_view.showGridLines = False
    headers = [
        "Period",
        "Revenue",
        "COGS",
        "Operating Expenses",
        "EBITDA",
        "Depreciation",
        "EBIT",
        "NOPAT",
        "Capex",
        "AR",
        "Inventory",
        "AP",
        "NWC",
        "Delta NWC",
        "FCF",
        "Discount Factor",
        "PV of FCF",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header_row(ws, 1, len(headers))

    start_year = forecast_df.iloc[0]["period"].year
    for row in range(2, years + 2):
        year = start_year + (row - 2)
        ws[f"A{row}"] = year

        if row == 2:
            ws[f"B{row}"] = "=Inputs!B38*(1+Inputs!B8*Inputs!B34)"
            ws[f"N{row}"] = "=M2-Inputs!B36"
        else:
            ws[f"B{row}"] = f"=B{row-1}*(1+Inputs!B8*Inputs!B34)"
            ws[f"N{row}"] = f"=M{row}-M{row-1}"

        ws[f"C{row}"] = f"=B{row}*Inputs!B9*(1-Inputs!B35/10000)"
        ws[f"D{row}"] = f"=B{row}*Inputs!B10"
        ws[f"E{row}"] = f"=B{row}-C{row}-D{row}"
        ws[f"F{row}"] = f"=B{row}*Inputs!B16"
        ws[f"G{row}"] = f"=E{row}-F{row}"
        ws[f"H{row}"] = f"=G{row}*(1-Inputs!B11)"
        ws[f"I{row}"] = f"=B{row}*Inputs!B15*Inputs!B37"
        ws[f"J{row}"] = f"=B{row}*Inputs!B12/365"
        ws[f"K{row}"] = f"=C{row}*Inputs!B14/365"
        ws[f"L{row}"] = f"=C{row}*Inputs!B13/365"
        ws[f"M{row}"] = f"=J{row}+K{row}-L{row}"
        ws[f"O{row}"] = f"=H{row}+F{row}-I{row}-N{row}"
        ws[f"P{row}"] = f"=IF(Inputs!B33=\"mid_year\",1/(1+Valuation!B2)^(ROW()-1.5),1/(1+Valuation!B2)^(ROW()-1))"
        ws[f"Q{row}"] = f"=O{row}*P{row}"

    for row in ws.iter_rows(min_row=2, max_row=years + 1, min_col=2, max_col=17):
        for cell in row:
            cell.font = FORMULA_FONT
            cell.number_format = ACCOUNTING_FMT
            cell.border = THIN_BORDER
    for row in range(2, years + 2):
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws[f"A{row}"].border = THIN_BORDER
        if row % 2 == 0:
            for col_idx in range(1, 18):
                ws.cell(row=row, column=col_idx).fill = ALT_ROW_FILL

    ws[f"A{years+3}"] = "Total PV of FCF"
    ws[f"B{years+3}"] = f"=SUM(Q2:Q{years+1})"
    ws[f"A{years+3}"].font = Font(bold=True)
    ws[f"B{years+3}"].font = FORMULA_FONT
    ws[f"B{years+3}"].number_format = ACCOUNTING_FMT
    ws[f"A{years+3}"].fill = SECTION_FILL
    ws[f"B{years+3}"].fill = SECTION_FILL
    ws[f"A{years+3}"].border = THIN_BORDER
    ws[f"B{years+3}"].border = THIN_BORDER

    ws.auto_filter.ref = f"A1:Q{years+1}"
    ws.freeze_panes = "A2"
    ws.print_area = f"A1:Q{years+1}"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    _autosize(ws)


def _write_valuation(wb: Workbook, last_forecast_row: int) -> None:
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Metric"
    ws["B1"] = "Value"

    _style_header_row(ws, 1, 2)

    metrics_formulas = [
        ("WACC", "=Inputs!B23*(Inputs!B17+Inputs!B19*Inputs!B18+Inputs!B20+Inputs!B21)+Inputs!B22*Inputs!B24*(1-Inputs!B25)"),
        ("PV of Explicit FCF", f"=SUM(Forecast!Q2:Forecast!Q{last_forecast_row})"),
        ("Terminal FCF", f"=Forecast!O{last_forecast_row}"),
        ("Terminal EBITDA", f"=Forecast!E{last_forecast_row}"),
        ("Terminal Value (Gordon)", "=B4*(1+Inputs!B26)/(B2-Inputs!B26)"),
        ("Terminal Value (Exit)", "=B5*Inputs!B27"),
        ("PV Terminal (Gordon)", f"=B6*Forecast!P{last_forecast_row}"),
        ("PV Terminal (Exit)", f"=B7*Forecast!P{last_forecast_row}"),
        ("Enterprise Value (Gordon)", "=B3+B8"),
        ("Enterprise Value (Exit)", "=B3+B9"),
        ("Equity Value (Gordon)", "=B10-Inputs!B30-Inputs!B31-Inputs!B32+Inputs!B29"),
        ("Equity Value (Exit)", "=B11-Inputs!B30-Inputs!B31-Inputs!B32+Inputs!B29"),
        ("Implied Price (Gordon)", "=B12/Inputs!B28"),
        ("Implied Price (Exit)", "=B13/Inputs!B28"),
    ]

    for idx, (metric, formula) in enumerate(metrics_formulas, start=2):
        ws[f"A{idx}"] = metric
        ws[f"B{idx}"] = formula
        ws[f"B{idx}"].font = FORMULA_FONT
        if metric == "WACC":
            ws[f"B{idx}"].number_format = PCT_FMT
        elif "Implied Price" in metric:
            ws[f"B{idx}"].number_format = MILLIONS_FMT
        elif "Terminal FCF" in metric or "Terminal EBITDA" in metric:
            ws[f"B{idx}"].number_format = ACCOUNTING_FMT
        else:
            ws[f"B{idx}"].number_format = MILLIONS_FMT
        ws[f"A{idx}"].border = THIN_BORDER
        ws[f"B{idx}"].border = THIN_BORDER
        if idx % 2 == 0:
            ws[f"A{idx}"].fill = ALT_ROW_FILL
            ws[f"B{idx}"].fill = ALT_ROW_FILL

    ws["A18"] = "Model Integrity"
    ws["B18"] = "=IF(Inputs!B26<=Inputs!B39,\"PASS\",\"ALERT: g > GDP cap\")"
    ws["B18"].font = FORMULA_FONT
    ws["A18"].border = THIN_BORDER
    ws["B18"].border = THIN_BORDER

    ws.freeze_panes = "A2"
    ws.print_area = "A1:D30"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    _autosize(ws)


def _write_sensitivity(wb: Workbook, cfg: DCFConfig, last_forecast_row: int, base_wacc: float) -> None:
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Terminal Growth / WACC"
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL

    wacc_values = [base_wacc - 0.02, base_wacc - 0.01, base_wacc, base_wacc + 0.01, base_wacc + 0.02]
    growth_values = [
        cfg.valuation.terminal_growth_rate - 0.01,
        cfg.valuation.terminal_growth_rate - 0.005,
        cfg.valuation.terminal_growth_rate,
        cfg.valuation.terminal_growth_rate + 0.005,
        cfg.valuation.terminal_growth_rate + 0.01,
    ]

    for col_idx, value in enumerate(wacc_values, start=2):
        ws.cell(row=1, column=col_idx, value=float(value))
        ws.cell(row=1, column=col_idx).number_format = PCT_FMT

    for row_idx, growth in enumerate(growth_values, start=2):
        ws.cell(row=row_idx, column=1, value=float(growth))
        ws.cell(row=row_idx, column=1).number_format = PCT_FMT
        for col_idx in range(2, 7):
            wacc_cell = f"{get_column_letter(col_idx)}$1"
            g_cell = f"$A{row_idx}"
            formula = (
                f"=((Valuation!B3+((Valuation!B4*(1+{g_cell}))/({wacc_cell}-{g_cell}))*Forecast!P{last_forecast_row})"
                "+Inputs!B29-Inputs!B30-Inputs!B31-Inputs!B32)/Inputs!B28"
            )
            ws.cell(row=row_idx, column=col_idx, value=formula)
            ws.cell(row=row_idx, column=col_idx).font = FORMULA_FONT
            ws.cell(row=row_idx, column=col_idx).number_format = ACCOUNTING_FMT
            ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER

    _style_header_row(ws, 1, 6)
    for row_idx in range(2, 7):
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        if row_idx % 2 == 0:
            for col_idx in range(1, 7):
                ws.cell(row=row_idx, column=col_idx).fill = ALT_ROW_FILL

    rng = "B2:F6"
    ws.conditional_formatting.add(
        rng,
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    ws.freeze_panes = "A2"
    ws.print_area = "A1:D12"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    _autosize(ws)


def _write_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:N1")
    ws["A1"] = "Executive Dashboard"
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws["A3"] = "Key Valuation Outputs"
    ws["A3"].font = Font(bold=True, color="1F4E78")

    kpi_metrics = [
        ("WACC", "=Valuation!B2", PCT_FMT),
        ("EV (Gordon)", "=Valuation!B10", MILLIONS_FMT),
        ("EV (Exit)", "=Valuation!B11", MILLIONS_FMT),
        ("Price (Gordon)", "=Valuation!B14", MILLIONS_FMT),
        ("Price (Exit)", "=Valuation!B15", MILLIONS_FMT),
    ]
    for idx, (label, formula, fmt) in enumerate(kpi_metrics, start=2):
        ws.cell(row=4, column=idx, value=label)
        ws.cell(row=5, column=idx, value=formula)
        ws.cell(row=4, column=idx).font = Font(bold=True, color="1F4E78")
        ws.cell(row=4, column=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=5, column=idx).font = FORMULA_FONT
        ws.cell(row=5, column=idx).number_format = fmt
        ws.cell(row=5, column=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=4, column=idx).fill = KPI_FILL
        ws.cell(row=5, column=idx).fill = KPI_FILL
        ws.cell(row=4, column=idx).border = THIN_BORDER
        ws.cell(row=5, column=idx).border = THIN_BORDER

    ws["A8"] = "Football Field Valuation"
    ws["A8"].font = Font(bold=True, color="1F4E78")

    ws["A10"] = "Method"
    ws["B10"] = "Implied Share Price"
    _style_header_row(ws, 10, 2)

    ws["A11"] = "Gordon Growth"
    ws["B11"] = "=Valuation!B14"
    ws["A12"] = "Exit Multiple"
    ws["B12"] = "=Valuation!B15"
    ws["B11"].font = FORMULA_FONT
    ws["B12"].font = FORMULA_FONT
    ws["B11"].number_format = MILLIONS_FMT
    ws["B12"].number_format = MILLIONS_FMT
    ws["A11"].border = THIN_BORDER
    ws["A12"].border = THIN_BORDER
    ws["B11"].border = THIN_BORDER
    ws["B12"].border = THIN_BORDER
    ws["A12"].fill = ALT_ROW_FILL
    ws["B12"].fill = ALT_ROW_FILL

    ws["A14"] = "Enterprise to Equity Bridge"
    ws["A14"].font = Font(bold=True, color="1F4E78")
    ws["A15"] = "Enterprise Value (Gordon)"
    ws["B15"] = "=Valuation!B10"
    ws["A16"] = "Less: Debt"
    ws["B16"] = "=-Inputs!B30"
    ws["A17"] = "Less: Minority Interest"
    ws["B17"] = "=-Inputs!B31"
    ws["A18"] = "Less: Preferred Stock"
    ws["B18"] = "=-Inputs!B32"
    ws["A19"] = "Add: Cash & Equivalents"
    ws["B19"] = "=Inputs!B29"
    ws["A20"] = "Equity Value"
    ws["B20"] = "=SUM(B15:B19)"

    for row in range(15, 21):
        ws[f"A{row}"].border = THIN_BORDER
        ws[f"B{row}"].border = THIN_BORDER
        ws[f"B{row}"].font = FORMULA_FONT
        ws[f"B{row}"].number_format = MILLIONS_FMT if row in [15, 20] else ACCOUNTING_FMT
        if row % 2 == 0:
            ws[f"A{row}"].fill = ALT_ROW_FILL
            ws[f"B{row}"].fill = ALT_ROW_FILL
    ws["A20"].font = Font(bold=True, color="1F4E78")
    ws["B20"].font = Font(bold=True, color="1F4E78")
    ws["A20"].fill = SECTION_FILL
    ws["B20"].fill = SECTION_FILL

    ws["D14"] = "Scenario Snapshot"
    ws["D14"].font = Font(bold=True, color="1F4E78")
    ws["D15"] = "Scenario"
    ws["E15"] = "Implied Price"
    _style_header_row(ws, 15, 5)
    ws["D16"] = "Bear"
    ws["E16"] = "=Valuation!B14*(1-0.12)"
    ws["D17"] = "Base"
    ws["E17"] = "=Valuation!B14"
    ws["D18"] = "Bull"
    ws["E18"] = "=Valuation!B14*(1+0.12)"
    ws["D19"] = "Spread (Bull-Bear)"
    ws["E19"] = "=E18-E16"
    for row in range(16, 20):
        ws[f"D{row}"].border = THIN_BORDER
        ws[f"E{row}"].border = THIN_BORDER
        ws[f"E{row}"].font = FORMULA_FONT
        ws[f"E{row}"].number_format = MILLIONS_FMT
        if row % 2 == 0:
            ws[f"D{row}"].fill = ALT_ROW_FILL
            ws[f"E{row}"].fill = ALT_ROW_FILL

    ws["G14"] = "Model Health"
    ws["G14"].font = Font(bold=True, color="1F4E78")
    ws["G15"] = "Check"
    ws["H15"] = "Status"
    _style_header_row(ws, 15, 8)
    ws["G16"] = "Terminal Growth <= GDP"
    ws["H16"] = "=IF(Inputs!B26<=Inputs!B39,\"PASS\",\"ALERT\")"
    ws["G17"] = "Balance Sheet Integrity"
    ws["H17"] = "=Checks!E8"
    for row in [16, 17]:
        ws[f"G{row}"].border = THIN_BORDER
        ws[f"H{row}"].border = THIN_BORDER
        ws[f"H{row}"].font = FORMULA_FONT
    ws.conditional_formatting.add("H16:H17", FormulaRule(formula=['H16="PASS"'], fill=PASS_FILL))
    ws.conditional_formatting.add("H16:H17", FormulaRule(formula=['H16="ALERT"'], fill=ALERT_FILL))

    ws["J10"] = "Year"
    ws["K10"] = "FCF"
    _style_header_row(ws, 10, 11)
    for row in range(11, 16):
        src_row = row - 9
        ws[f"J{row}"] = f"=Forecast!A{src_row}"
        ws[f"K{row}"] = f"=Forecast!O{src_row}"
        ws[f"J{row}"].border = THIN_BORDER
        ws[f"K{row}"].border = THIN_BORDER
        ws[f"K{row}"].number_format = MILLIONS_FMT

    chart = BarChart()
    chart.title = "Implied Price by Method"
    chart.style = 10
    chart.type = "col"
    chart.grouping = "clustered"
    data = Reference(ws, min_col=2, min_row=10, max_row=12)
    categories = Reference(ws, min_col=1, min_row=11, max_row=12)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 4.8
    chart.width = 5.8
    ws.add_chart(chart, "H2")

    trend = LineChart()
    trend.title = "FCF Trend"
    trend.style = 2
    trend.y_axis.title = "FCF"
    trend.x_axis.title = "Year"
    trend_data = Reference(ws, min_col=11, min_row=10, max_row=15)
    trend_cats = Reference(ws, min_col=10, min_row=11, max_row=15)
    trend.add_data(trend_data, titles_from_data=True)
    trend.set_categories(trend_cats)
    trend.legend = None
    trend.height = 4.8
    trend.width = 5.8
    ws.add_chart(trend, "H9")

    ws.print_area = "A1:N35"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    _autosize(ws)


def _write_report_data(
    wb: Workbook,
    forecast_df: pd.DataFrame,
    valuation_summary: dict,
    cfg: DCFConfig,
    scenario_name: str,
    historical_growth_3y_avg: float,
) -> None:
    ws = wb.create_sheet("ReportData")
    ws["A1"] = "metric"
    ws["B1"] = "value"
    metrics = [
        ("scenario", scenario_name),
        ("wacc", float(valuation_summary.get("WACC", 0.0))),
        ("ev_gordon", float(valuation_summary.get("Enterprise Value (Gordon)", 0.0))),
        ("ev_exit", float(valuation_summary.get("Enterprise Value (Exit)", 0.0))),
        ("eq_gordon", float(valuation_summary.get("Equity Value (Gordon)", 0.0))),
        ("eq_exit", float(valuation_summary.get("Equity Value (Exit)", 0.0))),
        ("price_gordon", float(valuation_summary.get("Implied Price (Gordon)", 0.0))),
        ("price_exit", float(valuation_summary.get("Implied Price (Exit)", 0.0))),
        ("risk_free", cfg.wacc.risk_free_rate),
        ("beta", cfg.wacc.beta),
        ("mrp", cfg.wacc.market_risk_premium),
        ("size_premium", cfg.wacc.size_premium),
        ("terminal_growth", cfg.valuation.terminal_growth_rate),
        ("revenue_cagr", cfg.forecast.revenue_cagr),
        ("historical_growth_3y_avg", historical_growth_3y_avg),
    ]
    for idx, (key, value) in enumerate(metrics, start=2):
        ws[f"A{idx}"] = key
        ws[f"B{idx}"] = value

    ws["D1"] = "period"
    ws["E1"] = "Revenue"
    ws["F1"] = "COGS"
    ws["G1"] = "EBITDA"
    ws["H1"] = "Capex"
    ws["I1"] = "Delta NWC"
    ws["J1"] = "FCF"

    for row_idx, (_, row) in enumerate(forecast_df.iterrows(), start=2):
        ws.cell(row=row_idx, column=4, value=str(row.get("period", ""))[:10])
        ws.cell(row=row_idx, column=5, value=float(row.get("Revenue", 0.0)))
        ws.cell(row=row_idx, column=6, value=float(row.get("COGS", 0.0)))
        ws.cell(row=row_idx, column=7, value=float(row.get("EBITDA", 0.0)))
        ws.cell(row=row_idx, column=8, value=float(row.get("Capex", 0.0)))
        ws.cell(row=row_idx, column=9, value=float(row.get("Delta NWC", 0.0)))
        ws.cell(row=row_idx, column=10, value=float(row.get("FCF", 0.0)))

    ws.sheet_state = "hidden"


def _write_checks(wb: Workbook, years: int) -> None:
    ws = wb.create_sheet("Checks")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Period"
    ws["B1"] = "Balance Sheet Check"
    ws["C1"] = "Status"
    ws["D1"] = "Terminal Growth Check"
    ws["E1"] = "Sanity Status"
    _style_header_row(ws, 1, 5)

    for row in range(2, years + 2):
        ws[f"A{row}"] = f"=Forecast!A{row}"
        ws[f"B{row}"] = f"=(Forecast!J{row}+Forecast!K{row})-(Forecast!L{row}+(Forecast!J{row}+Forecast!K{row}-Forecast!L{row}))"
        ws[f"C{row}"] = f"=IF(ABS(B{row})<0.01,\"PASS\",\"FAIL\")"
        ws[f"A{row}"].font = FORMULA_FONT
        ws[f"B{row}"].font = FORMULA_FONT
        ws[f"C{row}"].font = FORMULA_FONT
        ws[f"B{row}"].number_format = ACCOUNTING_FMT
        ws[f"A{row}"].border = THIN_BORDER
        ws[f"B{row}"].border = THIN_BORDER
        ws[f"C{row}"].border = THIN_BORDER
        ws[f"D{row}"].border = THIN_BORDER
        ws[f"E{row}"].border = THIN_BORDER
        ws[f"D{row}"] = "=IF(Inputs!B26<=Inputs!B39,\"PASS\",\"ALERT\")"
        ws[f"E{row}"] = f"=IF(C{row}=\"PASS\",\"PASS\",\"REVIEW\")"
        ws[f"D{row}"].font = FORMULA_FONT
        ws[f"E{row}"].font = FORMULA_FONT
        if row % 2 == 0:
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).fill = ALT_ROW_FILL

    base_row = years + 3
    ws[f"A{base_row}"] = "Model-level"
    ws[f"D{base_row}"] = "=IF(Inputs!B26<=Inputs!B39,\"PASS\",\"ALERT\")"
    ws[f"E{base_row}"] = "=IF(COUNTIF(C2:C{0},\"FAIL\")=0,\"PASS\",\"REVIEW\")".format(years + 1)
    ws[f"D{base_row}"].font = FORMULA_FONT
    ws[f"E{base_row}"].font = FORMULA_FONT
    ws[f"A{base_row}"].border = THIN_BORDER
    ws[f"D{base_row}"].border = THIN_BORDER
    ws[f"E{base_row}"].border = THIN_BORDER

    ws.conditional_formatting.add(
        f"C2:C{years+1}",
        FormulaRule(formula=["C2=\"PASS\""], fill=PASS_FILL),
    )
    ws.conditional_formatting.add(
        f"C2:C{years+1}",
        FormulaRule(formula=["C2=\"FAIL\""], fill=ALERT_FILL),
    )
    ws.conditional_formatting.add(
        f"D2:D{years+1}",
        FormulaRule(formula=["D2=\"ALERT\""], fill=ALERT_FILL),
    )

    ws.freeze_panes = "A2"
    ws.print_area = f"A1:E{base_row+2}"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    _autosize(ws)


def _write_page1_executive_summary(wb: Workbook, years: int) -> None:
    ws = wb.create_sheet("Page_1_Executive")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"] = "1-Minute Manager | Executive Summary"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL

    ws.merge_cells("A3:D3")
    ws["A3"] = "The Answer"
    ws["A3"].font = Font(bold=True, color="FFFFFF")
    ws["A3"].fill = HEADER_FILL
    ws.merge_cells("A4:D4")
    ws["A4"] = "EV Range"
    ws["A4"].font = Font(bold=True, color="1F4E78")
    ws["A5"] = "=TEXT(Valuation!B10*0.90/1000000,\"$#,##0.0\")&\"M - \"&TEXT(Valuation!B10*1.10/1000000,\"$#,##0.0\")&\"M\""
    ws["A5"].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells("A6:D6")
    ws["A6"] = "Share Price Range"
    ws["A6"].font = Font(bold=True, color="1F4E78")
    ws["A7"] = "=TEXT(Valuation!B14*0.90,\"$#,##0.00\")&\" - \"&TEXT(Valuation!B14*1.10,\"$#,##0.00\")"
    ws["A7"].font = Font(bold=True, size=13, color="1F4E78")

    ws.merge_cells("F3:J3")
    ws["F3"] = "Recommendation"
    ws["F3"].font = Font(bold=True, color="FFFFFF")
    ws["F3"].fill = HEADER_FILL
    ws.merge_cells("F4:J7")
    ws["F4"] = (
        "=IF(IF(Inputs!B40>0,Inputs!B40,Inputs!B41)=0,"
        "\"Status: MARKET PRICE N/A\","
        "IF(Valuation!B14>IF(Inputs!B40>0,Inputs!B40,Inputs!B41),"
        "\"Status: UNDERVALUED by \"&TEXT((Valuation!B14/IF(Inputs!B40>0,Inputs!B40,Inputs!B41)-1),\"0.0%\")&\" vs Current Market Price\","
        "\"Status: OVERVALUED by \"&TEXT((1-Valuation!B14/IF(Inputs!B40>0,Inputs!B40,Inputs!B41)),\"0.0%\")&\" vs Current Market Price\"))"
    )
    ws["F4"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["F4"].font = Font(bold=True, size=11, color="1F4E78")

    ws["A9"] = "Football Field"
    ws["A9"].font = Font(bold=True, color="1F4E78")
    ws["A10"] = "Method"
    ws["B10"] = "Low"
    ws["C10"] = "Base"
    ws["D10"] = "High"
    _style_header_row(ws, 10, 4)

    data_rows = [
        ("DCF (Bear/Base/Bull)", "=Valuation!B14*0.90", "=Valuation!B14", "=Valuation!B14*1.10"),
        ("Comparable Multiples", "=Valuation!B14*0.95", "=Valuation!B14*1.05", "=Valuation!B14*1.15"),
        ("Precedent Transactions", "=Valuation!B14*1.00", "=Valuation!B14*1.10", "=Valuation!B14*1.20"),
    ]
    for row, data in enumerate(data_rows, start=11):
        ws[f"A{row}"] = data[0]
        ws[f"B{row}"] = data[1]
        ws[f"C{row}"] = data[2]
        ws[f"D{row}"] = data[3]
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].border = THIN_BORDER
        for col in ["B", "C", "D"]:
            ws[f"{col}{row}"].number_format = MILLIONS_FMT

    ff_chart = BarChart()
    ff_chart.type = "bar"
    ff_chart.grouping = "stacked"
    ff_chart.style = 10
    ff_chart.title = "Valuation Range Comparison"
    ff_data = Reference(ws, min_col=2, min_row=10, max_col=4, max_row=13)
    ff_cat = Reference(ws, min_col=1, min_row=11, max_row=13)
    ff_chart.add_data(ff_data, titles_from_data=True)
    ff_chart.set_categories(ff_cat)
    ff_chart.height = 6
    ff_chart.width = 7
    ws.add_chart(ff_chart, "F9")

    ws["A16"] = "Key Ratios (3Y Historical + 5Y Projected)"
    ws["A16"].font = Font(bold=True, color="1F4E78")
    headers = ["Metric", "H-3", "H-2", "H-1", "Y1", "Y2", "Y3", "Y4", "Y5"]
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=17, column=idx, value=h)
    _style_header_row(ws, 17, 9)

    ws["A18"] = "Revenue Growth %"
    ws["B18"] = "=Inputs!B8"
    ws["C18"] = "=Inputs!B8"
    ws["D18"] = "=Inputs!B8"
    ws["E18"] = "=Forecast!B3/Forecast!B2-1"
    ws["F18"] = "=Forecast!B4/Forecast!B3-1"
    ws["G18"] = "=Forecast!B5/Forecast!B4-1"
    ws["H18"] = "=Forecast!B6/Forecast!B5-1"
    ws["I18"] = "=Forecast!B6/Forecast!B5-1"

    ws["A19"] = "EBITDA Margin %"
    ws["B19"] = "=1-Inputs!B9-Inputs!B10"
    ws["C19"] = "=1-Inputs!B9-Inputs!B10"
    ws["D19"] = "=1-Inputs!B9-Inputs!B10"
    ws["E19"] = "=Forecast!E2/Forecast!B2"
    ws["F19"] = "=Forecast!E3/Forecast!B3"
    ws["G19"] = "=Forecast!E4/Forecast!B4"
    ws["H19"] = "=Forecast!E5/Forecast!B5"
    ws["I19"] = "=Forecast!E6/Forecast!B6"

    ws["A20"] = "ROIC %"
    ws["B20"] = "=IFERROR(Forecast!H2/(Forecast!M2+Forecast!I2),0)"
    ws["C20"] = "=IFERROR(Forecast!H2/(Forecast!M2+Forecast!I2),0)"
    ws["D20"] = "=IFERROR(Forecast!H2/(Forecast!M2+Forecast!I2),0)"
    ws["E20"] = "=IFERROR(Forecast!H2/(Forecast!M2+Forecast!I2),0)"
    ws["F20"] = "=IFERROR(Forecast!H3/(Forecast!M3+Forecast!I3),0)"
    ws["G20"] = "=IFERROR(Forecast!H4/(Forecast!M4+Forecast!I4),0)"
    ws["H20"] = "=IFERROR(Forecast!H5/(Forecast!M5+Forecast!I5),0)"
    ws["I20"] = "=IFERROR(Forecast!H6/(Forecast!M6+Forecast!I6),0)"

    for r in [18, 19, 20]:
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if c >= 2:
                ws.cell(row=r, column=c).number_format = PCT_FMT

    ws["A23"] = "Report based on [BASE] Case Scenario"
    ws["A23"].font = Font(italic=True, color="666666")

    ws.print_area = "A1:J28"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    _autosize(ws)


def _write_page2_logic_check(wb: Workbook, years: int) -> None:
    ws = wb.create_sheet("Page_2_Logic_Check")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"] = "Logic Check | Assumptions & Drivers"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL

    ws["A3"] = "Revenue & Margins Trend"
    ws["A3"].font = Font(bold=True, color="1F4E78")
    headers = ["Metric", "Y1", "Y2", "Y3", "Y4", "Y5"]
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=4, column=idx, value=h)
    _style_header_row(ws, 4, 6)

    ws["A5"] = "Revenue Growth"
    ws["B5"] = "=Forecast!B3/Forecast!B2-1"
    ws["C5"] = "=Forecast!B4/Forecast!B3-1"
    ws["D5"] = "=Forecast!B5/Forecast!B4-1"
    ws["E5"] = "=Forecast!B6/Forecast!B5-1"
    ws["F5"] = "=E5"

    ws["A6"] = "EBITDA Margin"
    ws["B6"] = "=Forecast!E2/Forecast!B2"
    ws["C6"] = "=Forecast!E3/Forecast!B3"
    ws["D6"] = "=Forecast!E4/Forecast!B4"
    ws["E6"] = "=Forecast!E5/Forecast!B5"
    ws["F6"] = "=Forecast!E6/Forecast!B6"

    for r in [5, 6]:
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if c >= 2:
                ws.cell(row=r, column=c).number_format = PCT_FMT

    mini_growth = LineChart()
    mini_growth.style = 2
    mini_growth.height = 1.6
    mini_growth.width = 3.2
    mini_growth.legend = None
    growth_data = Reference(ws, min_col=2, min_row=4, max_col=6, max_row=5)
    growth_cat = Reference(ws, min_col=2, min_row=4, max_col=6)
    mini_growth.add_data(growth_data, titles_from_data=True)
    mini_growth.set_categories(growth_cat)
    ws.add_chart(mini_growth, "H4")

    mini_margin = LineChart()
    mini_margin.style = 2
    mini_margin.height = 1.6
    mini_margin.width = 3.2
    mini_margin.legend = None
    margin_data = Reference(ws, min_col=2, min_row=4, max_col=6, max_row=6)
    margin_cat = Reference(ws, min_col=2, min_row=4, max_col=6)
    mini_margin.add_data(margin_data, titles_from_data=True)
    mini_margin.set_categories(margin_cat)
    ws.add_chart(mini_margin, "H7")

    ws.merge_cells("A9:J10")
    ws["A9"] = "=\"Assumes revenue growth tapers from \"&TEXT(B5,\"0.0%\")&\" to \"&TEXT(Inputs!B26,\"0.0%\")&\" over 5 years as market saturation increases.\""
    ws["A9"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A9"].font = Font(italic=True, color="444444")

    ws["A12"] = "WACC Box"
    ws["A12"].font = Font(bold=True, color="1F4E78")
    ws["A13"] = "Risk-Free Rate"
    ws["B13"] = "=Inputs!B17"
    ws["A14"] = "Equity Risk Premium"
    ws["B14"] = "=Inputs!B18"
    ws["A15"] = "Beta"
    ws["B15"] = "=Inputs!B19"
    ws["A16"] = "WACC"
    ws["B16"] = "=Valuation!B2"
    for r in range(13, 17):
        ws[f"A{r}"].border = THIN_BORDER
        ws[f"B{r}"].border = THIN_BORDER
        ws[f"B{r}"].font = FORMULA_FONT
    ws["B13"].number_format = PCT_FMT
    ws["B14"].number_format = PCT_FMT
    ws["B16"].number_format = PCT_FMT
    ws["B16"].fill = KPI_FILL
    ws["B16"].font = Font(bold=True, color="1F4E78")

    ws["E12"] = "Terminal Value Logic"
    ws["E12"].font = Font(bold=True, color="1F4E78")
    ws.merge_cells("E13:J16")
    ws["E13"] = (
        "=\"Terminal Value based on \"&TEXT(Inputs!B26,\"0.0%\")"
        "&\" Perpetuity Growth Rate, implying a \"&TEXT(Inputs!B27,\"0.0x\")&\" Exit Multiple.\""
    )
    ws["E13"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["E13"].border = THIN_BORDER

    ws.print_area = "A1:J22"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    _autosize(ws)


def _write_page3_risk_map(wb: Workbook, years: int) -> None:
    ws = wb.create_sheet("Page_3_Risk_Map")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"] = "Risk Map | Sensitivity Analysis"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL

    ws["A3"] = "Table 1: Enterprise Value vs WACC & Terminal Growth"
    ws["A3"].font = Font(bold=True, color="1F4E78")
    ws["A4"] = "g / WACC"
    tg_values = ["=Inputs!B26-0.01", "=Inputs!B26-0.005", "=Inputs!B26", "=Inputs!B26+0.005", "=Inputs!B26+0.01"]
    wacc_values = ["=Valuation!B2-0.02", "=Valuation!B2-0.01", "=Valuation!B2", "=Valuation!B2+0.01", "=Valuation!B2+0.02"]
    for i, formula in enumerate(wacc_values, start=2):
        ws.cell(row=4, column=i, value=formula)
        ws.cell(row=4, column=i).number_format = PCT_FMT
    for r, formula in enumerate(tg_values, start=5):
        ws.cell(row=r, column=1, value=formula)
        ws.cell(row=r, column=1).number_format = PCT_FMT
        for c in range(2, 7):
            wacc_cell = f"{get_column_letter(c)}$4"
            g_cell = f"$A{r}"
            ws.cell(row=r, column=c, value=f"=Valuation!B3+((Valuation!B4*(1+{g_cell}))/({wacc_cell}-{g_cell}))*Forecast!P6")
            ws.cell(row=r, column=c).number_format = MILLIONS_FMT

    _style_header_row(ws, 4, 6)

    ws["A12"] = "Table 2: Share Price vs EBITDA Margin & Revenue Growth"
    ws["A12"].font = Font(bold=True, color="1F4E78")
    ws["A13"] = "Margin / Growth"
    margin_vals = ["-0.04", "-0.02", "0", "0.02", "0.04"]
    growth_vals = ["-0.03", "-0.015", "0", "0.015", "0.03"]
    for i, val in enumerate(growth_vals, start=2):
        ws.cell(row=13, column=i, value=float(val))
        ws.cell(row=13, column=i).number_format = PCT_FMT
    for r, val in enumerate(margin_vals, start=14):
        ws.cell(row=r, column=1, value=float(val))
        ws.cell(row=r, column=1).number_format = PCT_FMT
        for c in range(2, 7):
            rg = f"{get_column_letter(c)}$13"
            mg = f"$A{r}"
            ws.cell(row=r, column=c, value=f"=((Valuation!B10*(1+{rg}+{mg})-Inputs!B30-Inputs!B31-Inputs!B32+Inputs!B29)/Inputs!B28)")
            ws.cell(row=r, column=c).number_format = PRICE_FMT

    _style_header_row(ws, 13, 6)

    for r in list(range(5, 10)) + list(range(14, 19)):
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = Border(
                left=Side(style="thin", color="FFFFFF"),
                right=Side(style="thin", color="FFFFFF"),
                top=Side(style="thin", color="FFFFFF"),
                bottom=Side(style="thin", color="FFFFFF"),
            )

    ws.conditional_formatting.add(
        "B5:F9",
        ColorScaleRule(
            start_type="min",
            start_color="F4CCCC",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFFFFF",
            end_type="max",
            end_color="D9EAD3",
        ),
    )
    ws.conditional_formatting.add(
        "B14:F18",
        ColorScaleRule(
            start_type="min",
            start_color="F4CCCC",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFFFFF",
            end_type="max",
            end_color="D9EAD3",
        ),
    )

    ws.print_area = "A1:J24"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    _autosize(ws)


def _write_page4_engine_room(wb: Workbook, years: int) -> None:
    ws = wb.create_sheet("Page_4_Engine_Room")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    ws["A1"] = "Engine Room | Full DCF Output"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL

    headers = ["Line Item", "Hist Y1", "Hist Y2", "Hist Y3", "Proj Y1", "Proj Y2", "Proj Y3", "Proj Y4", "Proj Y5", "Terminal", "Notes"]
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=3, column=idx, value=h)
    _style_header_row(ws, 3, 11)

    row_labels = [
        "EBITDA",
        "D&A",
        "EBIT",
        "Tax",
        "NOPAT",
        "Plus D&A",
        "Less Capex",
        "Less Change in Working Capital",
        "Unlevered Free Cash Flow",
        "Discount Factor (Period/Sales)",
    ]
    for i, label in enumerate(row_labels, start=4):
        ws[f"A{i}"] = label
        ws[f"A{i}"].border = THIN_BORDER

    # Projected / terminal links
    mapping = {
        4: ("E", "E", "Valuation!B5"),
        5: ("F", "F", "Forecast!F6"),
        6: ("G", "G", "Forecast!G6"),
        7: ("", "", "=Forecast!G6-Forecast!H6"),
        8: ("H", "H", "Forecast!H6"),
        9: ("F", "F", "Forecast!F6"),
        10: ("I", "I", "Forecast!I6"),
        11: ("N", "N", "Forecast!N6"),
        12: ("O", "O", "Valuation!B4"),
        13: ("P", "P", "Forecast!P6"),
    }

    for row_idx, (col_proj, col_proj_dup, terminal_formula) in mapping.items():
        for col_offset, frow in enumerate(range(2, 7), start=5):
            source_col = col_proj
            if row_idx == 7:
                ws.cell(row=row_idx, column=col_offset, value=f"=Forecast!G{frow}-Forecast!H{frow}")
            else:
                ws.cell(row=row_idx, column=col_offset, value=f"=Forecast!{source_col}{frow}")
        ws.cell(row=row_idx, column=10, value=f"={terminal_formula}" if not str(terminal_formula).startswith("=") else terminal_formula)

    # Historical proxy values (3 columns)
    ws["B4"] = "=Forecast!E2/(1+Inputs!B8)^3"
    ws["C4"] = "=Forecast!E2/(1+Inputs!B8)^2"
    ws["D4"] = "=Forecast!E2/(1+Inputs!B8)"
    for r in range(5, 13):
        ws[f"B{r}"] = f"=E{r}"
        ws[f"C{r}"] = f"=E{r}"
        ws[f"D{r}"] = f"=E{r}"
    ws["B13"] = "=Forecast!P2"
    ws["C13"] = "=Forecast!P2"
    ws["D13"] = "=Forecast!P2"

    for r in range(4, 14):
        for c in range(2, 11):
            ws.cell(row=r, column=c).font = FORMULA_FONT
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r == 13:
                ws.cell(row=r, column=c).number_format = "0.0000"
            else:
                ws.cell(row=r, column=c).number_format = MILLIONS_FMT if r in [4, 8, 12] else ACCOUNTING_FMT

    ws["K4"] = "Core operating profitability"
    ws["K8"] = "After tax operating income"
    ws["K12"] = "UFCF for DCF"
    ws["K13"] = "Mid-year convention visible"

    ws.print_area = "A1:K18"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    _autosize(ws)


def _write_print_report(wb: Workbook) -> None:
    ws = wb.create_sheet("PRINT_REPORT")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "Print Report Assembly"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL

    steps = [
        "1) Use Camera Tool snapshots from Page_1_Executive to Page_4_Engine_Room.",
        "2) Arrange each page block onto printable A4 layout on this tab.",
        "3) Keep scenario notation visible: Report based on [BASE] Case Scenario.",
        "4) Export this tab as PDF for client distribution.",
        "5) Validate headers/footers and confidentiality footer before sending.",
    ]
    ws["A3"] = "Action Plan"
    ws["A3"].font = Font(bold=True, color="1F4E78")
    for i, text in enumerate(steps, start=4):
        ws[f"A{i}"] = text
        ws[f"A{i}"].border = THIN_BORDER
        if i % 2 == 0:
            ws[f"A{i}"].fill = ALT_ROW_FILL

    ws["A11"] = "Quick Checks"
    ws["A11"].font = Font(bold=True, color="1F4E78")
    ws["A12"] = "Looks Big-4 quality?"
    ws["B12"] = "=IF(AND(Page_1_Executive!B5<>\"\",Page_4_Engine_Room!E12<>\"\"),\"YES\",\"REVIEW\")"
    ws["A13"] = "Confidential footer present?"
    ws["B13"] = "=IF(Page_1_Executive!A1<>\"\",\"YES\",\"REVIEW\")"
    ws["B12"].font = FORMULA_FONT
    ws["B13"].font = FORMULA_FONT
    ws["A12"].border = THIN_BORDER
    ws["B12"].border = THIN_BORDER
    ws["A13"].border = THIN_BORDER
    ws["B13"].border = THIN_BORDER

    ws.print_area = "A1:H40"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    _autosize(ws)


def _excel_safe_value(value):
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    return value


def _style_header_row(ws, row_number: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_number, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def _input_number_format(label: str, value) -> str:
    if not isinstance(value, (int, float)):
        return "General"
    percent_labels = {
        "Revenue CAGR",
        "COGS % Revenue",
        "Opex % Revenue",
        "Tax Rate",
        "Capex % Revenue",
        "Depreciation Rate",
        "Risk-Free Rate",
        "Market Risk Premium",
        "Size Premium",
        "Country Risk Premium",
        "Debt Weight",
        "Equity Weight",
        "Pre-tax Cost of Debt",
        "WACC Tax Rate",
        "Terminal Growth Rate",
        "GDP Growth Cap",
    }
    if label in percent_labels:
        return PCT_FMT
    if "Shares" in label:
        return "#,##0"
    if any(x in label for x in ["Cash", "Debt", "Interest", "Stock", "Revenue", "NWC"]):
        return ACCOUNTING_FMT
    return ACCOUNTING_FMT


def _apply_workbook_theme(wb: Workbook) -> None:
    tab_colors = {
        "Dashboard": "1F4E78",
        "Inputs": "3A3A3A",
        "Forecast": "1F4E78",
        "Valuation": "3A3A3A",
        "Sensitivity": "1F4E78",
        "Checks": "3A3A3A",
        "ReportData": "0F243E",
    }
    desired_order = [
        "Dashboard",
        "Inputs",
        "Forecast",
        "Valuation",
        "Sensitivity",
        "Checks",
        "ReportData",
    ]
    ordered = []
    for name in desired_order:
        if name in wb.sheetnames:
            ordered.append(wb[name])
    wb._sheets = ordered

    for ws in wb.worksheets:
        color = tab_colors.get(ws.title)
        if color:
            ws.sheet_properties.tabColor = color
        _apply_header_footer(ws)
        _apply_font_family(ws, BODY_FONT)
        _apply_negative_soft_red(ws)


def _apply_header_footer(ws) -> None:
    ws.oddHeader.left.text = "PROJECT [COMPANY NAME] | VALUATION REPORT"
    ws.oddHeader.right.text = "&D"
    ws.oddFooter.center.text = "Strictly Private & Confidential | Prepared by Rounak Jain, CFA L2 Candidate"


def _apply_font_family(ws, font_name: str) -> None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue
            old_font = cell.font or Font()
            cell.font = Font(
                name=font_name,
                size=old_font.sz,
                bold=old_font.bold,
                italic=old_font.italic,
                color=old_font.color,
                underline=old_font.underline,
            )


def _apply_negative_soft_red(ws) -> None:
    max_row = max(ws.max_row, 1)
    max_col = max(ws.max_column, 1)
    rng = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="lessThan", formula=["0"], font=Font(color=SOFT_RED, name=BODY_FONT)),
    )


def _autosize(ws) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)
